import subprocess
import os

def install_dependencies(pip_command):
    if os.path.isfile('dep_inst.txt'): # let's check if dependancies are already installed
        return
    
    try:
        print("\033[3;2mInstalling dependencies...\033[0m", end=" ")
        subprocess.check_call([pip_command, "install", "-r", "requirements.txt", "--disable-pip-version-check"], stdout=subprocess.PIPE, stderr=subprocess.DEVNULL)
        print("\033[3;2mDependencies installed successfully!\033[0m")
        with open('dep_inst.txt', 'w') as f:
            f.write('True') # write a flag to indicate that the dependencies are installed
    except subprocess.CalledProcessError:
        print("\033[3;31mFailed to install dependencies.\033[0m")

def determine_pip_command():
    """
    Determine which version of pip to use.
    """
    devnull = open(os.devnull, 'w')
    pip_command = None
    if subprocess.call(["pip", "--version"], stdout=devnull, stderr=devnull) == 0:
        pip_command = "pip"
    elif subprocess.call(["pip3", "--version"], stdout=devnull, stderr=devnull) == 0:
        pip_command = "pip3"
    else:
        print("\033[3;31mError: pip is not installed.\033[0m")
        exit()
    return pip_command

#check the dependancies and install them
install_dependencies(determine_pip_command())


import csv
import pandas as pd
from termcolor import colored
from tqdm import tqdm
import openpyxl


def calculate_elo_rating(winner, loser, winner_score, loser_score):
    K = 50  # constant factor to adjust the weight of each game


    # Compute the expected win probabilities for each player on the winning and losing teams
    winner_expected = 1 / (1 + 10 ** ((loser[0] + loser[1] - winner[0] - winner[1]) / 400))
    loser_expected = 1 / (1 + 10 ** ((winner[0] + winner[1] - loser[0] - loser[1]) / 400))
   

    # Determine the point differential and calculate the corresponding adjustment factor
    point_diff = abs(winner_score - loser_score)
    if point_diff == 5:
        adj_factor = 1
    else:
        adj_factor = 0.2 * point_diff

    # Calculate the elo rating changes for each player on the winning and losing teams
    winner_delta = round(K * adj_factor * (1 - winner_expected))
    loser_delta = round(K * adj_factor * (0 - loser_expected))

    # Update the elo scores for each player on the winning and losing teams
    new_winner_elo = (winner[0] + winner_delta, winner[1] + winner_delta)
    new_loser_elo = (loser[0] + loser_delta, loser[1] + loser_delta)

    # Return the new elo scores for both teams
    return (new_winner_elo, new_loser_elo)


def update_players(players, team1, team2, score1, score2):
    """
    Updates the players' ELO ratings based on match results and returns updated players dictionary.
    """
    winner1, winner2 = team1 if score1 > score2 else team2
    loser1, loser2 = team2 if score1 > score2 else team1
    new_winner_elo, new_loser_elo = calculate_elo_rating((players[winner1],players[winner2]), (players[loser1],players[loser2]), score1,score2)
    
    players[winner1] = new_winner_elo[0] if new_winner_elo[0] > 0 else 0
    players[winner2] = new_winner_elo[1] if new_winner_elo[1] > 0 else 0
    players[loser1] = new_loser_elo[0] if new_loser_elo[0] > 0 else 0
    players[loser2] = new_loser_elo[1] if new_loser_elo[1] > 0 else 0
    
    # Write point history of each player to the excel file
    for player in [winner1, winner2, loser1, loser2]:
        filename = "player_stats/player_stats.xlsx"
        sheetname = f"{player}"
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        if sheetname not in workbook.sheetnames:
            workbook.create_sheet(sheetname)
        sheet = workbook[sheetname]
        sheet.append([players[player]])
        workbook.save(filename)
    
    
    return players


def print_ranking(players):
    """
    Prints the player rankings in the order of ELO rating, from highest to lowest.
    """
    ranking = sorted(players.items(), key=lambda x: x[1], reverse=True)
    headers = ["Rank", "Player", "Points"]
    rows = []
    for i, (player, points) in enumerate(ranking, start=1):
        rank = f"{i}"
        if i == 1:
            player = f"*{player}"
        rows.append([rank, player, round(points)])
    
    # Set the width of each column
    col_widths = [6, 25, 8]
    
    # Print the headers
    print(f"\n\033[1m{headers[0]:<{col_widths[0]}}{headers[1]:<{col_widths[1]}}{headers[2]:<{col_widths[2]}}\033[0m")
    
    # Print the separator
    print("=" * sum(col_widths))
    
    # Print the rows
    for row in rows:
        print(f"{row[0]:<{col_widths[0]}}{row[1]:<{col_widths[1]}}{row[2]:<{col_widths[2]}}")
    
    # Create a pandas dataframe from the rankings
    df = pd.DataFrame(rows, columns=headers)
    print("\033[3;2m\nExporting to ranks.xlsx...\033[0m",end=" ")
    # Write the dataframe to an Excel file
    df.to_excel('ranks.xlsx', index=False)
    print("\033[3;2mdone!\033[0m")


def read_csv_file(file_path):
    rows = []
    with open(file_path, newline='') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=',', quotechar='"')
        next(csvreader) # skip the first row
        for row in csvreader:
            if not all(row):  # skip rows with missing values
                continue
            rows.append(row)
    return rows

def main():
    players = {}  # initialize an empty dictionary to store player names and ELO ratings

# Create directory if it doesn't exist
    if not os.path.exists('player_stats'):
        os.makedirs('player_stats')
    if os.path.exists('player_stats/player_stats.xlsx'):
        os.remove('player_stats/player_stats.xlsx') # delete old stats to prevent duplications

    reader = read_csv_file('scores.csv')
    for row in tqdm(reader, desc="Processing..."):
        if len(row) >= 6:
            team1 = row[0], row[1]
            team2 = row[2], row[3]
            score1, score2 = int(row[4]), int(row[5])
            for player in team1 + team2:
                if player not in players:
                    players[player] = int(1000)  # initialize new player's score to 1000
            players = update_players(players, team1, team2, score1, score2)
    
    print_ranking(players)


# run the program
if __name__ == "__main__":
    main() 
